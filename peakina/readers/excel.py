"""
Module to add excel files support
"""
import datetime
import logging
from functools import wraps
from typing import Any, Dict, Generator, List, Optional, Set, Tuple, Union

import openpyxl
import pandas as pd
import xlrd

LOGGER = logging.getLogger(__name__)


def _old_xls_rows_iterator(
    wb: Union[openpyxl.workbook.Workbook, xlrd.book.Book],
    sh_name: str,
    preview_nrows: Optional[int],
    preview_offset: Optional[int],
) -> Generator[Any, Any, Any]:
    """
    Depending on paginations inputs (preview_rows, preview_offset), we want to
    get an iterator object to loop on target rows, here we're returning an iterator
    using yield for each iteration in the workbook

    """

    if preview_nrows is None and preview_offset is not None:
        to_iter = range(preview_offset, wb.sheet_by_name(sh_name).nrows)
    elif preview_nrows is not None and preview_offset is not None:
        to_iter = range(preview_offset, preview_offset + preview_nrows + 1)
    elif preview_nrows is not None and preview_offset is None:
        to_iter = range(preview_nrows + 1)
    else:
        to_iter = range(wb.sheet_by_name(sh_name).nrows)

    for rx in to_iter:
        try:
            yield wb.sheet_by_name(sh_name).row(rx)
        except IndexError:
            break


def _new_xls_rows_iterator(
    wb: Union[openpyxl.workbook.Workbook, xlrd.book.Book],
    sh_name: str,
    preview_nrows: Optional[int],
    preview_offset: Optional[int],
) -> Generator[Any, Any, Any]:
    """
    Depending on paginations inputs (preview_rows, preview_offset), we want to
    get an iterator object to loop on target rows, here we're returning an iterator
    from the iter_rows built-in function from openpyxl

    """

    # +1 are here because this is 1-based indexing
    if preview_nrows is not None and preview_offset is not None:
        max_row = preview_offset + 1 + preview_nrows
    elif preview_nrows is not None and preview_offset is None:
        max_row = preview_nrows + 1
    else:
        max_row = None

    if preview_offset:
        min_row = preview_offset + 1
    else:
        min_row = None

    # Then we return the generator
    yield wb[sh_name].iter_rows(
        min_row=min_row,
        max_row=max_row,
        values_only=True,
    )


def _get_rows_iterator(
    wb: Union[openpyxl.workbook.Workbook, xlrd.book.Book],
    sheet_name: str,
    preview_nrows: Optional[int],
    preview_offset: Optional[int],
) -> Generator[Any, Any, Any]:
    """
    Depending on the excel type either it's the new format or the old one,
    this method will return an iterator to read on its rows
    """

    if isinstance(wb, xlrd.book.Book):
        return _old_xls_rows_iterator(wb, sheet_name, preview_nrows, preview_offset)

    return _new_xls_rows_iterator(wb, sheet_name, preview_nrows, preview_offset)


def _build_row_subset(
    row: Union[List[Any], Tuple[Any]],
    sh_name: str,
    sheetnames: List[str],
    row_number: int,
    row_subset: List[str],
) -> Tuple[List[str], Set[int]]:
    """
    This method will build each row and add an extra row for the sheet_name
    If we're in an excel with multiple sheets

    """
    date_columns_indices = set()

    def _infer_type(cell_value: Any, cell_index: int) -> Any:
        value = str(cell_value)
        if type(cell_value) in [int, float, str]:
            # we're removing "," from cells because we're going to be using comma as seperator for our csv payload
            # and if we keep some cells with comma, it could generate fake mismatch errors on columns...
            value = str(cell_value).replace(",", " ")
        elif type(cell_value) == bool:
            # we're assuming "True" and "False" will be considered as booleans
            value = f'"{cell_value}"'
        elif type(cell_value) in [datetime.datetime]:
            # in teh context of only preview, i think it's okay to
            # just have a representation of the date
            value = cell_value.isoformat()
            date_columns_indices.add(cell_index)
        return value

    cells = [
        _infer_type(cell.value, cell_index)
        if type(cell) not in [str, int, float, bool, datetime.datetime] and cell is not None
        else _infer_type(cell, cell_index)
        for cell_index, cell in enumerate(row)
    ]

    if len(sheetnames) > 1:
        row_subset.append(f'{",".join([*cells, sh_name])}\n')
    else:
        row_subset.append(f'{",".join(cells)}\n')

    return row_subset, date_columns_indices


def _get_row_subset_per_sheet(
    wb: Union[openpyxl.workbook.Workbook, xlrd.book.Book],
    sh_name: str,
    sheetnames: List[str],
    preview_nrows: Optional[int],
    preview_offset: Optional[int],
    row_subset: List[str],
    skiprows: Optional[int] = None,
    nrows: Optional[int] = None,
    skipfooter: int = 0,
) -> Tuple[List[str], Set[int]]:
    """
    This method will get an iterator from the workbook and
    construct a list of row inside row_subset
    """
    # we get the row iterator from here
    row_iterator = _get_rows_iterator(wb, sh_name, preview_nrows, preview_offset)
    date_columns_indices: Set[int] = set()

    def __loop_and_fill_row_subsets(
        row_subset: List[str], loop_on: Any
    ) -> Tuple[List[str], Set[int]]:
        headers_skipped = False
        date_columns_indices: Set[int] = set()
        for row_number, row in loop_on:
            # We want to skip the headers if we're in another sheet
            if not headers_skipped:
                headers_skipped = True
                continue
            if skiprows:
                if row_number <= skiprows:
                    continue
            row_subset, date_columns_indices = _build_row_subset(
                row, sh_name, sheetnames, row_number, row_subset
            )
            if nrows:
                if row_number == nrows:
                    break

        return row_subset, date_columns_indices

    if isinstance(wb, openpyxl.workbook.Workbook):
        for row_iter in row_iterator:
            row_subset, date_columns_indices = __loop_and_fill_row_subsets(
                row_subset, enumerate(row_iter)
            )
    else:
        row_subset, date_columns_indices = __loop_and_fill_row_subsets(
            row_subset, enumerate(row_iterator)
        )

    # to handle the skipfooter
    lines_to_keep = len(row_subset) - skipfooter

    return row_subset[:lines_to_keep], date_columns_indices


def _read_sheets(
    wb: Union[openpyxl.workbook.Workbook, xlrd.book.Book],
    sheet_names: List[Any],
    preview_nrows: Optional[int],
    preview_offset: Optional[int],
    nrows: Optional[int] = None,
    skiprows: Optional[int] = None,
    skipfooter: int = 0,
) -> Tuple[List[Any], Set[int]]:
    """
    This method will loop over sheets, read content and return a list of rows
    depending on your inputs

    """

    row_subset: List[str] = []
    date_columns_indices: Set[int] = set()
    for sh_name in sheet_names:
        row_subset, date_columns_indices = _get_row_subset_per_sheet(
            wb,
            sh_name,
            sheet_names,
            preview_nrows,
            preview_offset,
            row_subset,
            skiprows,
            nrows,
            skipfooter,
        )

    if isinstance(wb, openpyxl.workbook.Workbook):
        wb.close()

    return row_subset, date_columns_indices


@wraps(pd.read_excel)
def read_excel(
    *args: Any,
    preview_nrows: Optional[int] = None,
    preview_offset: int = 0,
    **kwargs: Any,
) -> pd.DataFrame:

    df_or_dict: Union[Dict[str, pd.DataFrame], pd.DataFrame] = pd.read_excel(*args, **kwargs)

    if isinstance(df_or_dict, dict):  # multiple sheets
        for sheet_name, sheet_df in df_or_dict.items():
            sheet_df["__sheet__"] = sheet_name
        df = pd.concat(list(df_or_dict.values()))
    else:
        df = df_or_dict

    if preview_nrows:
        return df[preview_offset : preview_offset + preview_nrows or 0]

    return df


def excel_meta(filepath: str, reader_kwargs: Dict[str, Any]) -> Dict[str, Any]:
    """
    Returns a dictionary with the meta information of the excel file.
    """
    excel_file = pd.ExcelFile(filepath)
    sheet_names = excel_file.sheet_names

    df = read_excel(excel_file, **reader_kwargs)

    if (sheet_name := reader_kwargs.get("sheet_name", 0)) is None:
        # multiple sheets together
        return {
            "sheetnames": sheet_names,
            "df_rows": df.shape[0],
            "total_rows": sum(excel_file.parse(sheet_name).shape[0] for sheet_name in sheet_names),
        }
    else:
        # single sheet
        return {
            "sheetnames": sheet_names,
            "df_rows": df.shape[0],
            "total_rows": excel_file.parse(sheet_name).shape[0],
        }
